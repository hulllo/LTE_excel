from openpyxl.reader.excel import load_workbook
class Excel():
    def __init__(self, filename):
        self.filename = filename
    def writeto(self, list_): 
        wb=load_workbook(self.filename)
        sheetnames  = wb.sheetnames
        for x in list_:
            sheet_name = x[0]
            row = x[1]
            col = x[2]
            value = x[3]
            if sheet_name not in sheetnames:
                print('sheet_name do not exist')
                return False
            else:
                ws = wb[sheet_name]
                ws.cell(row,col).value = value
        wb.save('./a_changed.xlsx')

def main():
    excel = Excel('./a.xlsx')
    excel.writeto('LTEB3', 12, 6, '23')
            
            
if __name__ == '__main__':
    main()
